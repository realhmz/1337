import tkinter as tk
from tkinter import messagebox, ttk
import openpyxl
from PIL import Image, ImageDraw, ImageFont, ImageTk
from bidi.algorithm import get_display

# Specify the path to an Arabic-supported TrueType font file
font_path = "/home/hamza/Desktop/fondation/hello.ttf"

# Load the font with a suitable size
font_size = 20
font = ImageFont.truetype(font_path, font_size, encoding='utf-8')

# Create an image with Arabic text
def create_arabic_text_image(text, width, height):
    image = Image.new('RGB', (width, height), 'white')
    draw = ImageDraw.Draw(image)
    if text:
        draw.text((5, -3), text, fill='black', font=font, direction="rtl")
    return image

def add_data():
    data = [data_entries[i].get() for i in range(len(column_names))]
    if all(data):
        try:
            wb = openpyxl.load_workbook('haf_data.xlsx')
            sheet = wb.active
            sheet.append(data)
            wb.save('haf_data.xlsx')
            for entry in data_entries:
                entry.delete(0, 'end')
            messagebox.showinfo('Success', 'Data added successfully!')
        except Exception as e:
            messagebox.showerror('Error', str(e))
    else:
        messagebox.showerror('Error', 'Please enter data for all columns.')

# Load column names from the Excel file
def load_column_names():
    try:
        wb = openpyxl.load_workbook('haf_data.xlsx')
        sheet = wb.active
        for row in sheet.iter_rows(min_row=1, max_row=1):
            for cell in row:
                column_names.append(cell.value)
    except Exception as e:
        messagebox.showerror('Error', 'Failed to load column names. Please ensure "haf_data.xlsx" exists and is formatted correctly.')
        root.destroy()

def on_canvas_scroll(event):
    canvas.yview_scroll(-1 * (event.delta // 120), "units")

# Create the main window
root = tk.Tk()
root.title('Excel Data Entry')
root.geometry('1500x600')  # Lowered window size by 20 pixels in height

# Define custom colors
background_color = '#F5F5F5'
entry_color = '#ECF0F1'
button_color = '#E74C3C'

column_names = []
load_column_names()

data_entries = []

# Define custom colors for each category
category_colors = ['#3498db', '#27ae60', '#e74c3c']

# Create a canvas with a scroll bar
canvas = tk.Canvas(root)
canvas.pack(side='left', fill='both', expand=True)

scrollbar = tk.Scrollbar(root, orient='vertical', command=canvas.yview)
scrollbar.pack(side='right', fill='y')

canvas.configure(yscrollcommand=scrollbar.set)
canvas.bind_all("<MouseWheel>", on_canvas_scroll)  # Bind mouse wheel scrolling

frame = tk.Frame(canvas, bg=background_color)
canvas.create_window((0, 0), window=frame, anchor='nw')

# Create data entry fields and labels for each column with category-specific colors
label_entry_pairs = []  # To store label-entry pairs

# Define the number of names and data boxes in each column
column_counts = [10, 6, len(column_names) - 10 - 6]

# Define category titles
category_titles = ["ﺔﻳﻮﻬﻟﺍ تﺎﻣﻮﻠﻌﻣ", "ﺔﻴﻋﻮﻨﻟﺍ تﺎﻣﻮﻠﻌﻣ", "ﺔﻴﻤﻜﻟﺍ تﺎﻣﻮﻠﻌﻣ"]

for col_index, count in enumerate(column_counts):
    # Add category title label
    title_label = tk.Label(frame, text=category_titles[col_index], font=('Arial', 16, 'bold'), fg='white', bg=category_colors[col_index])
    title_label.grid(row=0, column=0 + col_index * 2, columnspan=2, padx=10, pady=10, sticky='nsew')

    for i in range(count):
        entry = tk.Entry(frame, bg=entry_color)
        entry.grid(row=i + 1, column=1 + col_index * 2, padx=10, pady=10, sticky='nsew')
        data_entries.append(entry)

        label_text = column_names[i + sum(column_counts[:col_index])] if col_index > 0 else column_names[i]
        label_image = create_arabic_text_image(label_text, 200, 30)
        label_image = ImageTk.PhotoImage(label_image)  # Convert to PhotoImage

        label = tk.Label(frame, image=label_image, font=('Arial', 14), fg='white', bg=category_colors[col_index])
        label.image = label_image  # Keep a reference to the image to prevent it from being garbage collected
        label.grid(row=i + 1, column=0 + col_index * 2, padx=10, pady=10, sticky='nsew')

        label_entry_pairs.append((label, entry))

# Create and configure add data button using ttk.Style
style = ttk.Style()
style.configure("TButton", background=button_color, foreground='white')
add_button = ttk.Button(root, text='Add Data', style="TButton", command=add_data, cursor='hand2')
add_button.pack(side='top', padx=20, pady=10)

# Create a search entry and button
search_entry = tk.Entry(root, font=('Arial', 14), bg=entry_color)
search_entry.pack(side='left', padx=20, pady=10)

def search_data():
    query = search_entry.get().strip()
    if not query:
        messagebox.showerror('Error', 'Please enter a search query.')
        return

    try:
        wb = openpyxl.load_workbook('haf_data.xlsx')
        sheet = wb.active
        search_result = []

        for row in sheet.iter_rows(min_row=2):  # Assuming the data starts from row 2
            for cell in row:
                if query in str(cell.value):
                    search_result_row = []
                    for cell in row:
                        # Apply get_display to Arabic text
                        bidi_text = get_display(str(cell.value))
                        search_result_row.append(bidi_text)
                    search_result.append(search_result_row)

        if search_result:
            show_search_results(search_result)
        else:
            messagebox.showinfo('Search Result', 'No matching data found.')
    except Exception as e:
        messagebox.showerror('Error', str(e))

def show_search_results(results):
    result_window = tk.Toplevel(root)
    result_window.title('Search Result')

    search_result_frame = tk.Frame(result_window)
    search_result_frame.pack(fill='both', expand=True)

    search_result_canvas = tk.Canvas(search_result_frame)
    search_result_canvas.pack(side='left', fill='both', expand=True)

    search_result_scrollbar = tk.Scrollbar(search_result_frame, orient='vertical', command=search_result_canvas.yview)
    search_result_scrollbar.pack(side='right', fill='y')

    search_result_canvas.configure(yscrollcommand=search_result_scrollbar.set)

    search_result_frame_inner = tk.Frame(search_result_canvas)
    search_result_canvas.create_window((0, 0), window=search_result_frame_inner, anchor='nw')

    for i, row in enumerate(results):
        for j, cell in enumerate(row):
            label = tk.Label(search_result_frame_inner, text=cell, font=('Arial', 14), padx=10, pady=5, relief='solid')
            label.grid(row=i, column=j, sticky='nsew')

    search_result_frame_inner.update_idletasks()
    search_result_canvas.configure(scrollregion=search_result_canvas.bbox('all'))

# Create a search button
search_button = ttk.Button(root, text='Search', style='TButton', command=search_data)
search_button.pack(side='left', padx=10, pady=10)

# Start the main loop
root.mainloop()

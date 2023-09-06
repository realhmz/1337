import tkinter as tk
from tkinter import messagebox, ttk
import openpyxl
from PIL import Image, ImageDraw, ImageFont, ImageTk

# Specify the path to an Arabic-supported TrueType font file
font_path = "/home/hamza/Desktop/fondation/simpo.ttf"

# Load the font with a suitable size
font_size = 20
font = ImageFont.truetype(font_path, font_size, encoding='utf-8')

# Create an image with Arabic text
def create_arabic_text_image(text, width, height):
    image = Image.new('RGB', (width, height), 'white')
    draw = ImageDraw.Draw(image)
    if text:
        draw.text((5, -1), text, fill='black', font=font, direction="rtl")
    return image

def add_data():
    data = [data_entries[i].get() for i in range(len(column_names))]
    if all(data):
        try:
            wb = openpyxl.load_workbook('haf_data.xlsx')
            sheet = wb.active
            sheet.append(data)
            wb.save('haf_data.xlsx')
            for entry in data_entries:
                entry.delete(0, 'end')
            messagebox.showinfo('Success', 'Data added successfully!')
        except Exception as e:
            messagebox.showerror('Error', str(e))
    else:
        messagebox.showerror('Error', 'Please enter data for all columns.')

# Load column names from the Excel file
def load_column_names():
    try:
        wb = openpyxl.load_workbook('haf_data.xlsx')
        sheet = wb.active
        for row in sheet.iter_rows(min_row=1, max_row=1):
            for cell in row:
                column_names.append(cell.value)
    except Exception as e:
        messagebox.showerror('Error', 'Failed to load column names. Please ensure "haf_data.xlsx" exists and is formatted correctly.')
        root.destroy()

def on_canvas_scroll(event):
    canvas.yview_scroll(-1 * (event.delta // 120), "units")

# Create the main window
root = tk.Tk()
root.title('Excel Data Entry')
root.geometry('1024x768')  # Set a fixed window size

column_names = []
load_column_names()

data_entries = []

# Create a canvas with a scroll bar
canvas = tk.Canvas(root)
canvas.pack(side='left', fill='both', expand=True)

scrollbar = tk.Scrollbar(root, orient='vertical', command=canvas.yview)
scrollbar.pack(side='right', fill='y')

canvas.configure(yscrollcommand=scrollbar.set)
canvas.bind_all("<MouseWheel>", on_canvas_scroll)  # Bind mouse wheel scrolling

frame = tk.Frame(canvas)
canvas.create_window((0, 0), window=frame, anchor='nw')

# Create a list to store label-entry pairs
label_entry_pairs = []

# Create data entry fields and labels for each column
for i, column_name in enumerate(column_names):
    col = 0 if i < 10 else 1 if i < 16 else 2
    row = i % 10 if col == 0 else (i - 10) % 6 if col == 1 else (i - 16)

    entry = tk.Entry(frame)
    entry.grid(row=row, column=col * 2 + 1, padx=10, pady=10, sticky='nsew')
    data_entries.append(entry)

    label_text = column_name
    label_image = create_arabic_text_image(label_text, 200, 30)
    label_image = ImageTk.PhotoImage(label_image)  # Convert to PhotoImage

    label = tk.Label(frame, image=label_image, font=('Arial', 14), fg='black')
    label.image = label_image  # Keep a reference to the image to prevent it from being garbage collected
    label.grid(row=row, column=col * 2, padx=10, pady=10, sticky='nsew')

    label_entry_pairs.append((label, entry))

# Create and configure add data button
add_button = ttk.Button(root, text='Add Data', style='TButton', command=add_data)
add_button.pack(side='top', padx=20, pady=10)

# Calculate the canvas scroll region to show all elements
frame.update_idletasks()
canvas.config(scrollregion=canvas.bbox('all'))

# Start the main loop
root.mainloop()
